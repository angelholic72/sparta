from openpyxl import load_workbook

work_book = load_workbook('prac01.xlsx')
work_sheet = work_book['prac']

# 데이터를 읽어봅니다.

work_sheet.cell(row=2, column=2, value='홍길동')

work_book.save('prac01.xlsx')





movie = list(db.movies.find({'star':sound_star}))

for i in movie:
    print(i['title'])
